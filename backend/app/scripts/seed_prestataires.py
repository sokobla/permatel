"""
Seeding PERMATEL — Import Prestataires ADM → Clients & Sites.

Lit le fichier Excel `Prestataires ADM.xlsx` (feuilles `Prestataires` et
`Sites`) et crée ou met à jour les entités correspondantes dans PERMATEL,
scoppées sur le tenant cible.

Commande Flask associée (enregistrée dans app/__init__.py) :

  flask seed-prestataires --tenant-code <CODE> [options]

Options :
  --tenant-code   CODE    Tenant cible (obligatoire).
  --file          PATH    Chemin vers le fichier Excel (défaut : chemin embarqué).
  --dry-run               Simule sans écriture (défaut : activé).
  --no-dry-run            Applique les écritures.
  --verbose               Affiche le détail de chaque enregistrement.
  --yes                   Bypass la confirmation interactive.

Arbitrages appliqués :
  1. code_site   = slug normalisé depuis le nom du site.
  2. Code postal  = float → str zero-padded sur 5 chiffres (auto).
  3. Ville        = conservée telle quelle, CEDEX inclus.
  4. Noms         = normalisés en Title Case.
  5. Tenant       = spécifié via --tenant-code.
  6. Ré-exécution = UPDATE des champs si l'entité existe déjà.
  7. Téléphone    = valeur '-' → NULL.
"""

import os
import re
import unicodedata
from pathlib import Path

import click
from flask.cli import with_appcontext
from sqlalchemy.exc import IntegrityError

from app import db

# ── Chemin par défaut vers le fichier source ─────────────────────────────────
_HERE = Path(__file__).resolve().parent          # app/scripts/
_DEFAULT_FILE = (
    _HERE.parent.parent / "resources" / "Prestataires ADM.xlsx"
)


# ─────────────────────────────────────────────────────────────────────────────
#  Utilitaires de normalisation
# ─────────────────────────────────────────────────────────────────────────────

def _fix_str(value) -> str | None:
    """Retourne une chaîne nettoyée ou None si vide / None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _title(value) -> str | None:
    """Normalise une chaîne en Title Case après nettoyage."""
    s = _fix_str(value)
    if s is None:
        return None
    return s.title()


def _normalize_cp(value) -> str | None:
    """
    Convertit un code postal Excel (float ou str) en str à 5 chiffres.
    Ex : 59650.0 → '59650' | 1300.0 → '01300' | 6700.0 → '06700'.
    """
    if value is None:
        return None
    try:
        cp = str(int(float(value))).strip()
        return cp.zfill(5)
    except (ValueError, TypeError):
        return _fix_str(value)


def _normalize_phone(value) -> str | None:
    """Retourne None si la valeur est vide, '-' ou équivalent."""
    s = _fix_str(value)
    if s is None or s == "-":
        return None
    return s


def _slugify(value: str) -> str:
    """
    Transforme un nom en slug alphanumérique majuscule (50 car. max).
    Ex : 'CORA METZ TECHNOPOLE' → 'CORA_METZ_TECHNOPOLE'
         'ZOO D\'AMNEVILLE'     → 'ZOO_D_AMNEVILLE'
    """
    # Décomposer les caractères Unicode pour supprimer les accents
    nfkd = unicodedata.normalize("NFKD", value)
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    # Remplacer tout ce qui n'est pas alphanumérique par _
    slug = re.sub(r"[^A-Za-z0-9]+", "_", ascii_str)
    # Supprimer les _ de début/fin et passer en majuscules
    slug = slug.strip("_").upper()
    return slug[:50]


def _unique_slug(base: str, used: set) -> str:
    """
    Garantit l'unicité d'un slug dans le batch en ajoutant un suffixe
    numérique si nécessaire.
    Ex : 'CORA_METZ' déjà présent → 'CORA_METZ_2'.
    """
    candidate = base
    counter = 2
    while candidate in used:
        candidate = f"{base[:47]}_{counter}"
        counter += 1
    used.add(candidate)
    return candidate


# ─────────────────────────────────────────────────────────────────────────────
#  Lecture & normalisation du fichier Excel
# ─────────────────────────────────────────────────────────────────────────────

def _load_excel(filepath: Path, verbose: bool) -> tuple[list[dict], list[dict]]:
    """
    Charge et normalise les deux feuilles du fichier Excel.

    Retourne :
        prestataires : list[dict]  — données clients normalisées
        sites        : list[dict]  — données sites normalisées (avec code_site)

    Lève :
        FileNotFoundError  si le fichier est absent.
        ValueError         si une feuille est manquante.
        RuntimeError       si openpyxl n'est pas installé.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "La bibliothèque 'openpyxl' est requise : pip install openpyxl"
        )

    if not filepath.exists():
        raise FileNotFoundError(f"Fichier introuvable : {filepath}")

    wb = openpyxl.load_workbook(filepath)

    for sheet in ("Prestataires", "Sites"):
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"Feuille '{sheet}' absente du fichier. Feuilles disponibles : "
                f"{wb.sheetnames}"
            )

    # ── Feuille Prestataires ─────────────────────────────────────────────────
    ws_p = wb["Prestataires"]
    rows_p = list(ws_p.iter_rows(values_only=True))
    # Ligne 1 = en-têtes : ('code', 'Adresse', 'Ville', 'Code postal', 'telephone', 'Email', 'Nom')
    prestataires = []
    seen_codes = set()

    for i, row in enumerate(rows_p[1:], start=2):
        code, adresse, ville, cp, tel, email, nom = row

        code = _fix_str(code)
        if not code:
            click.echo(f"  WARN  Prestataires L{i} : colonne 'code' vide — ligne ignorée.", err=True)
            continue
        if code in seen_codes:
            click.echo(f"  WARN  Prestataires L{i} : code '{code}' dupliqué dans le fichier — ligne ignorée.", err=True)
            continue
        seen_codes.add(code)

        nom_norm = _title(nom)
        if not nom_norm:
            click.echo(f"  ERR   Prestataires L{i} ({code}) : nom vide — ligne ignorée.", err=True)
            continue

        prestataires.append({
            "code_client": code,
            "nom": nom_norm,
            "adresse": _fix_str(adresse),
            "ville": _title(ville),
            "code_postal": _normalize_cp(cp),
            "telephone": _normalize_phone(tel),
            "email": _fix_str(email).lower() if _fix_str(email) else None,
        })

        if verbose:
            click.echo(f"  [P]   {code} | {nom_norm}")

    # ── Feuille Sites ────────────────────────────────────────────────────────
    ws_s = wb["Sites"]
    rows_s = list(ws_s.iter_rows(values_only=True))
    # Ligne 1 = en-têtes : ('Nom', 'Adresse', 'ville', 'code postal', 'telephone', 'code prestataire')
    sites = []
    used_slugs: set[str] = set()
    prest_codes = {p["code_client"] for p in prestataires}

    for i, row in enumerate(rows_s[1:], start=2):
        nom, adresse, ville, cp, tel, code_prest = row

        nom_norm = _title(nom)
        if not nom_norm:
            click.echo(f"  ERR   Sites L{i} : nom vide — ligne ignorée.", err=True)
            continue

        code_prest = _fix_str(code_prest)
        if not code_prest:
            click.echo(f"  ERR   Sites L{i} ({nom_norm}) : 'code prestataire' vide — ligne ignorée.", err=True)
            continue
        if code_prest not in prest_codes:
            click.echo(
                f"  ERR   Sites L{i} ({nom_norm}) : code prestataire '{code_prest}' "
                f"inconnu dans la feuille Prestataires — ligne ignorée.", err=True
            )
            continue

        # Génération du code_site unique dans le batch
        base_slug = _slugify(nom_norm)
        code_site = _unique_slug(base_slug, used_slugs)

        sites.append({
            "nom": nom_norm,
            "code_site": code_site,
            "adresse": _fix_str(adresse),
            "ville": _fix_str(ville),           # conservé tel quel (CEDEX inclus)
            "code_postal": _normalize_cp(cp),
            "telephone": _normalize_phone(tel),
            "code_prestataire": code_prest,     # clé de jointure — résolu en Phase 3
        })

        if verbose:
            click.echo(f"  [S]   {code_site} | {nom_norm} → {code_prest}")

    return prestataires, sites


# ─────────────────────────────────────────────────────────────────────────────
#  Logique de seeding
# ─────────────────────────────────────────────────────────────────────────────

def run_seed(tenant_code: str, filepath: Path, dry_run: bool, verbose: bool) -> dict:
    """
    Exécute le pipeline de seeding complet.

    Retourne un dict de compteurs :
        {
          'clients': {'created': int, 'updated': int, 'rejected': int},
          'sites':   {'created': int, 'updated': int, 'rejected': int},
        }
    """
    from app.models.tenant import Tenant
    from app.models.client import Client
    from app.models.site import Site

    sep = "=" * 62

    click.echo(sep)
    click.echo("  PERMATEL — Seeding Prestataires ADM → Clients & Sites")
    click.echo(sep)

    # ── PHASE 0 : Résolution du tenant ───────────────────────────────────────
    tenant = Tenant.query.filter_by(code=tenant_code).first()
    if not tenant:
        raise click.ClickException(f"Tenant '{tenant_code}' introuvable en base.")
    if not tenant.is_active:
        raise click.ClickException(f"Tenant '{tenant_code}' est inactif. Arrêt.")

    click.echo(f"  Tenant  : {tenant.nom} (code={tenant.code}, id={tenant.id})")
    click.echo(f"  Fichier : {filepath}")
    click.echo(f"  Mode    : {'DRY-RUN (aucune écriture)' if dry_run else 'ÉCRITURE RÉELLE'}")
    click.echo(sep)

    # ── PHASE 1 : Chargement & normalisation ─────────────────────────────────
    click.echo("[PHASE 1] Chargement et normalisation du fichier...")

    prestataires, sites = _load_excel(filepath, verbose)

    click.echo(
        f"  OK    {len(prestataires)} prestataire(s) chargé(s) | "
        f"{len(sites)} site(s) chargé(s)."
    )

    # Avertissements sur les CP anormaux
    for s in sites:
        if s["code_postal"] and len(s["code_postal"].lstrip("0")) <= 3:
            click.echo(
                f"  WARN  CP suspect après correction : '{s['code_postal']}' "
                f"pour le site '{s['nom']}' — à vérifier."
            )

    sites_sans_tel = sum(1 for s in sites if s["telephone"] is None)
    click.echo(f"  WARN  {sites_sans_tel} site(s) sans téléphone (valeur '-' → NULL).")

    # ── PHASE 2 : Validation structurelle ────────────────────────────────────
    click.echo("[PHASE 2] Validation structurelle...")

    errors = []
    for p in prestataires:
        if not p["nom"]:
            errors.append(f"  Client '{p['code_client']}' : nom manquant.")

    if errors:
        for e in errors:
            click.echo(f"  ERR   {e}", err=True)
        raise click.ClickException(
            f"{len(errors)} erreur(s) bloquante(s) détectée(s). Arrêt avant écriture."
        )

    click.echo("  OK    Aucune anomalie bloquante.")

    # ── PHASE 3 : Clients (CREATE / UPDATE) ──────────────────────────────────
    click.echo("[PHASE 3] Synchronisation des Clients...")

    counters = {
        "clients": {"created": 0, "updated": 0, "rejected": 0},
        "sites":   {"created": 0, "updated": 0, "rejected": 0},
    }

    # Map code_client → client_id (nécessaire en Phase 4)
    code_to_client: dict[str, Client] = {}

    for p in prestataires:
        try:
            existing: Client | None = Client.query.filter_by(
                tenant_id=tenant.id,
                code_client=p["code_client"],
            ).first()

            if existing:
                # UPDATE
                existing.nom         = p["nom"]
                existing.adresse     = p["adresse"]
                existing.ville       = p["ville"]
                existing.code_postal = p["code_postal"]
                existing.telephone   = p["telephone"]
                existing.email       = p["email"]
                existing.is_active   = True
                db.session.flush()
                code_to_client[p["code_client"]] = existing
                counters["clients"]["updated"] += 1
                action = "MAJ"
            else:
                # CREATE
                client = Client(
                    tenant_id=tenant.id,
                    code_client=p["code_client"],
                    nom=p["nom"],
                    adresse=p["adresse"],
                    ville=p["ville"],
                    code_postal=p["code_postal"],
                    telephone=p["telephone"],
                    email=p["email"],
                    is_active=True,
                )
                db.session.add(client)
                db.session.flush()
                code_to_client[p["code_client"]] = client
                counters["clients"]["created"] += 1
                action = "CREE"

            if verbose:
                click.echo(
                    f"  {action:<5} Client : {p['code_client']} | {p['nom']}"
                )

        except IntegrityError as exc:
            db.session.rollback()
            counters["clients"]["rejected"] += 1
            click.echo(
                f"  ERR   Client '{p['code_client']}' rejeté (IntegrityError) : {exc.orig}",
                err=True,
            )

    click.echo(
        f"  OK    Clients — "
        f"créés : {counters['clients']['created']} | "
        f"mis à jour : {counters['clients']['updated']} | "
        f"rejetés : {counters['clients']['rejected']}"
    )

    # ── PHASE 4 : Sites (CREATE / UPDATE) ────────────────────────────────────
    click.echo("[PHASE 4] Synchronisation des Sites...")

    for s in sites:
        parent_client = code_to_client.get(s["code_prestataire"])
        if parent_client is None:
            counters["sites"]["rejected"] += 1
            click.echo(
                f"  ERR   Site '{s['nom']}' rejeté : client parent "
                f"'{s['code_prestataire']}' absent en base (création échouée).",
                err=True,
            )
            continue

        try:
            existing: Site | None = Site.query.filter_by(
                tenant_id=tenant.id,
                code_site=s["code_site"],
            ).first()

            if existing:
                # UPDATE — le (tenant_id, client_id) composite est préservé
                existing.nom         = s["nom"]
                existing.adresse     = s["adresse"]
                existing.ville       = s["ville"]
                existing.code_postal = s["code_postal"]
                existing.telephone   = s["telephone"]
                existing.client_id   = parent_client.id
                existing.is_active   = True
                db.session.flush()
                counters["sites"]["updated"] += 1
                action = "MAJ"
            else:
                # CREATE — tenant_id identique au client (contrainte composite)
                site = Site(
                    tenant_id=tenant.id,
                    client_id=parent_client.id,
                    nom=s["nom"],
                    code_site=s["code_site"],
                    adresse=s["adresse"],
                    ville=s["ville"],
                    code_postal=s["code_postal"],
                    telephone=s["telephone"],
                    is_active=True,
                )
                db.session.add(site)
                db.session.flush()
                counters["sites"]["created"] += 1
                action = "CREE"

            if verbose:
                click.echo(
                    f"  {action:<5} Site  : {s['code_site']} | {s['nom']} "
                    f"→ {s['code_prestataire']}"
                )

        except IntegrityError as exc:
            db.session.rollback()
            counters["sites"]["rejected"] += 1
            click.echo(
                f"  ERR   Site '{s['nom']}' rejeté (IntegrityError) : {exc.orig}",
                err=True,
            )

    click.echo(
        f"  OK    Sites — "
        f"créés : {counters['sites']['created']} | "
        f"mis à jour : {counters['sites']['updated']} | "
        f"rejetés : {counters['sites']['rejected']}"
    )

    # ── PHASE 5 : Commit ou Rollback ─────────────────────────────────────────
    if dry_run:
        db.session.rollback()
        click.echo("[PHASE 5] DRY-RUN — rollback. Aucune écriture effectuée.")
    else:
        db.session.commit()
        click.echo("[PHASE 5] COMMIT — données enregistrées.")

    # ── PHASE 6 : Résumé ─────────────────────────────────────────────────────
    click.echo(sep)
    click.echo("  RÉSUMÉ")
    click.echo(
        f"  Clients : "
        f"{counters['clients']['created']} créé(s) | "
        f"{counters['clients']['updated']} mis à jour | "
        f"{counters['clients']['rejected']} rejeté(s)"
    )
    click.echo(
        f"  Sites   : "
        f"{counters['sites']['created']} créé(s) | "
        f"{counters['sites']['updated']} mis à jour | "
        f"{counters['sites']['rejected']} rejeté(s)"
    )
    if dry_run:
        click.echo("  → Aucune écriture (dry-run). Relancer avec --no-dry-run pour appliquer.")
    click.echo(sep)

    return counters


# ─────────────────────────────────────────────────────────────────────────────
#  Commande Flask CLI
# ─────────────────────────────────────────────────────────────────────────────

@click.command("seed-prestataires")
@click.option(
    "--tenant-code", required=True,
    help="Code du tenant cible (ex. ACME). Obligatoire.",
)
@click.option(
    "--file", "filepath",
    default=str(_DEFAULT_FILE),
    show_default=True,
    type=click.Path(dir_okay=False),
    help="Chemin vers le fichier Excel source.",
)
@click.option(
    "--dry-run/--no-dry-run", default=True, show_default=True,
    help="Simule sans ecriture (defaut : active).",
)
@click.option(
    "--verbose", is_flag=True, default=False,
    help="Affiche le detail de chaque enregistrement traite.",
)
@click.option(
    "--yes", is_flag=True, default=False,
    help="Bypass la confirmation interactive (mode CI/CD).",
)
@with_appcontext
def seed_prestataires_command(tenant_code, filepath, dry_run, verbose, yes):
    """
    Importe Prestataires ADM -> Clients & Sites pour un tenant donne.

    Idempotent : les entites existantes sont mises a jour, les nouvelles
    sont creees. En dry-run (defaut), aucune ecriture n'est effectuee.

    Exemples :

    \b
      # Simulation (defaut)
      flask seed-prestataires --tenant-code ACME

    \b
      # Application reelle avec confirmation bypassee
      flask seed-prestataires --tenant-code ACME --no-dry-run --yes

    \b
      # Application avec fichier alternatif et mode verbeux
      flask seed-prestataires --tenant-code ACME --no-dry-run --file /tmp/prest.xlsx --verbose
    """
    filepath = Path(filepath)

    # Confirmation interactive avant écriture réelle
    if not dry_run and not yes:
        click.confirm(
            f"\n⚠ Écriture réelle sur le tenant '{tenant_code}'. Continuer ?",
            abort=True,
        )

    try:
        run_seed(
            tenant_code=tenant_code,
            filepath=filepath,
            dry_run=dry_run,
            verbose=verbose,
        )
    except FileNotFoundError as exc:
        raise click.ClickException(str(exc))
    except ValueError as exc:
        raise click.ClickException(str(exc))
    except RuntimeError as exc:
        raise click.ClickException(str(exc))
    except Exception as exc:
        db.session.rollback()
        raise click.ClickException(
            f"Erreur inattendue — rollback effectué.\n  {type(exc).__name__}: {exc}"
        )
